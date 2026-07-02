"""
Smart Insurance Skill 0611 - Excel报告生成器

生成保单解读Excel报告，包含多个工作表
"""

import json
from typing import Dict, Any, List
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelExporter:
    """
    Excel报告生成器 - 生成保单解读Excel报告

    使用示例:
        exporter = ExcelExporter()
        exporter.generate(data, './outputs/report.xlsx')
    """

    def __init__(self):
        """初始化Excel导出器"""
        if not EXCEL_AVAILABLE:
            print("Warning: openpyxl not installed, Excel export will not work")

    def generate(self, data: Dict[str, Any], output_path: str) -> str:
        """
        生成Excel报告

        Args:
            data: 解读结果数据
            output_path: 输出路径

        Returns:
            输出文件路径
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        # 创建工作簿
        wb = openpyxl.Workbook()

        # 定义样式
        header_fill = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
        header_font = Font(bold=True, color='333333')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D4D4D4'),
            right=Side(style='thin', color='D4D4D4'),
            top=Side(style='thin', color='D4D4D4'),
            bottom=Side(style='thin', color='D4D4D4')
        )

        # 1. 保单基本信息表
        ws1 = wb.active
        ws1.title = '保单基本信息'

        basic_info = data.get('保单基本信息') or data.get('basic_policy_information') or data.get('policy_basic_info', {})

        # 写入表头
        ws1['A1'] = '字段'
        ws1['B1'] = '值'
        ws1['A1'].fill = header_fill
        ws1['A1'].font = header_font
        ws1['B1'].fill = header_fill
        ws1['B1'].font = header_font

        # 写入数据
        field_order = [
            ('投保单位名称', 'name_of_the_insuring_entity'),
            ('投保单号/保单号', 'policy_number'),
            ('保险期间', 'insurance_period'),
            ('等待期', 'waiting_period'),
            ('保险来源', 'insurance_source'),
            ('险种类别', 'insurance_category'),
            ('保单状态', 'policy_status')
        ]

        row = 2
        for cn_name, en_name in field_order:
            value = basic_info.get(cn_name) or basic_info.get(en_name) or ''
            if isinstance(value, dict) and 'value' in value:
                value = value.get('value', '')

            ws1[f'A{row}'] = cn_name
            ws1[f'B{row}'] = str(value) if value else ''
            ws1[f'A{row}'].border = thin_border
            ws1[f'B{row}'].border = thin_border
            row += 1

        # 设置列宽
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 80

        # 2. 保险责任表
        ws2 = wb.create_sheet('保险责任')

        insurance_liability = data.get('保险责任') or data.get('insurance_liability') or []
        if '方案信息' in data:
            for scheme_name, scheme_data in data['方案信息'].items():
                if '保险责任' in scheme_data:
                    insurance_liability.extend(scheme_data['保险责任'])

        # 写入表头
        headers = ['层级名称', '方案序号', '保险种类', '保险责任', '保额', '公用关系',
                   '免赔额', '限额', '赔付比例（有医保）', '赔付比例（无医保）',
                   '赔付内容（甲类）', '赔付内容（乙类）', '赔付内容（乙类诊疗）', '赔付内容（自费）',
                   '指定医院', '方案特约']

        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # 写入数据
        row = 2
        for item in insurance_liability:
            level_name = item.get('层级名称') or item.get('level_name') or ''
            scheme_number = item.get('方案序号') or item.get('scheme_serial_number') or ''
            insurance_type = item.get('保险种类') or item.get('types_of_insurance') or ''
            liability = item.get('保险责任') or item.get('insurance_liability') or ''
            amount = item.get('保额') or item.get('insurance_coverage_amount') or ''
            shared = item.get('公用关系') or item.get('shared_or_not') or ''

            # 免赔额
            deductible = self._format_deductible(item)
            # 限额
            limit = self._format_limit(item)

            ratio_with = item.get('赔付比例（有医保）') or item.get('claim_payment_ratio_with_medical_insurance') or ''
            ratio_without = item.get('赔付比例（无医保）') or item.get('claim_payment_ratio_without_medical_insurance') or ''

            coverage_a = item.get('赔付内容（甲类）') or item.get('claim_coverage_class_A') or ''
            coverage_b_drug = item.get('赔付内容（乙类药）') or item.get('claim_coverage_class_B_drugs') or ''
            coverage_b_medical = item.get('赔付内容（乙类诊疗）') or item.get('claim_coverage_class_B_medical_services') or ''
            coverage_self = item.get('赔付内容（自费）') or item.get('claim_coverage_self_funded') or ''

            hospital = item.get('指定医院') or item.get('designated_hospital') or ''
            agreement = item.get('方案特约') or item.get('scheme_special_agreement') or ''

            # 提取值
            def extract_value(val):
                if isinstance(val, dict) and 'value' in val:
                    return val.get('value', '')
                return str(val) if val else ''

            values = [
                extract_value(level_name),
                extract_value(scheme_number),
                extract_value(insurance_type),
                extract_value(liability),
                extract_value(amount),
                extract_value(shared),
                deductible,
                limit,
                extract_value(ratio_with),
                extract_value(ratio_without),
                extract_value(coverage_a),
                extract_value(coverage_b_drug),
                extract_value(coverage_b_medical),
                extract_value(coverage_self),
                extract_value(hospital),
                extract_value(agreement)
            ]

            for col, value in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border

            row += 1

        # 设置列宽
        column_widths = [60, 55, 100, 180, 75, 55, 90, 90, 70, 70, 90, 90, 180, 90, 180, 180]
        for col, width in enumerate(column_widths, 1):
            ws2.column_dimensions[get_column_letter(col)].width = width

        # 3. 团单特约表
        ws3 = wb.create_sheet('团单特约')

        group_agreement = data.get('团单特约') or data.get('group_agreement') or {}

        ws3['A1'] = '字段'
        ws3['B1'] = '值'
        ws3['A1'].fill = header_fill
        ws3['A1'].font = header_font
        ws3['B1'].fill = header_fill
        ws3['B1'].font = header_font

        field_order = [
            ('既往症', 'pre_existing_condition'),
            ('是否持卡', 'card_holding_status'),
            ('特需就诊', 'special_needs_outpatient_service'),
            ('职业类别', 'job_category'),
            ('特殊计算', 'special_calculation'),
            ('药量控制', 'medication_dosage_control'),
            ('门特/门慢的特殊规则', 'chronic_disease_outpatient_services')
        ]

        row = 2
        for cn_name, en_name in field_order:
            value = group_agreement.get(cn_name) or group_agreement.get(en_name) or ''
            if isinstance(value, dict) and 'value' in value:
                value = value.get('value', '')

            ws3[f'A{row}'] = cn_name
            ws3[f'B{row}'] = str(value) if value else ''
            ws3[f'A{row}'].border = thin_border
            ws3[f'B{row}'].border = thin_border
            row += 1

        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 80

        # 4. 个人特约表
        ws4 = wb.create_sheet('个人特约')

        personal_agreement = data.get('个人特约') or data.get('personal_agreement') or []

        headers = ['序号', '姓名', '身份证号', '个人特约']
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        row = 2
        for idx, item in enumerate(personal_agreement):
            name = item.get('姓名') or item.get('name') or ''
            id_card = item.get('身份证号') or item.get('id_card_number') or ''
            agreement = item.get('个人特约') or item.get('personal_agreement') or ''

            ws4.cell(row=row, column=1, value=idx + 1)
            ws4.cell(row=row, column=2, value=name)
            ws4.cell(row=row, column=3, value=id_card)
            ws4.cell(row=row, column=4, value=agreement)

            row += 1

        ws4.column_dimensions['A'].width = 10
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 20
        ws4.column_dimensions['D'].width = 60

        # 5. 其他特约表
        ws5 = wb.create_sheet('其他特约')

        other_agreement = data.get('其他特约') or data.get('other_agreement') or []

        headers = ['序号', '特约内容']
        for col, header in enumerate(headers, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for item in other_agreement:
            serial = item.get('序号') or item.get('serial_number') or ''
            content = item.get('特约内容') or item.get('special_agreement_content') or ''

            ws5.cell(row=row, column=1, value=serial)
            ws5.cell(row=row, column=2, value=content)

            row += 1

        ws5.column_dimensions['A'].width = 10
        ws5.column_dimensions['B'].width = 100

        # 保存文件
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

        return str(path)

    def _format_deductible(self, item: Dict) -> str:
        """格式化免赔额"""
        values = []

        per_claim = item.get('次免赔额') or item.get('per_claim_deductible')
        daily = item.get('日免赔额') or item.get('daily_deductible')
        annual = item.get('年免赔额') or item.get('annual_deductible')

        if per_claim:
            values.append(f'次: {per_claim}')
        if daily:
            values.append(f'日: {daily}')
        if annual:
            values.append(f'年: {annual}')

        return '\n'.join(values) if values else ''

    def _format_limit(self, item: Dict) -> str:
        """格式化限额"""
        values = []

        per_claim = item.get('次限额') or item.get('per_claim_limit')
        daily = item.get('日限额') or item.get('daily_limit')
        annual = item.get('年限额') or item.get('annual_limit')

        if per_claim:
            values.append(f'次: {per_claim}')
        if daily:
            values.append(f'日: {daily}')
        if annual:
            values.append(f'年: {annual}')

        return '\n'.join(values) if values else ''